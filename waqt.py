import streamlit as st
import pandas as pd
from io import BytesIO
from supabase import create_client
import google.generativeai as genai
from openpyxl import load_workbook
from difflib import get_close_matches
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import re


st.set_page_config(page_title="Excel Auto-Updater for Waqt", layout="wide")

def suggest_column_name(col_name, available_columns):
    matches = get_close_matches(col_name, available_columns, n=1, cutoff=0.6)
    return matches[0] if matches else None

# --- Load environment variables ---
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]

if not SUPABASE_URL or not SUPABASE_KEY or not GEMINI_API_KEY:
    st.error("❌ Missing Supabase or Gemini credentials.")
    st.stop()

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel("gemini-2.0-flash")

# --- Split by blank rows ---
def split_dataframe_by_blank_rows(df):
    split_indices = df[df.isnull().all(axis=1)].index.tolist()
    blocks = []
    start_idx = 0

    for idx in split_indices:
        block = df.iloc[start_idx:idx]
        if not block.dropna(how="all").empty:
            blocks.append((start_idx, block.reset_index(drop=True)))
        start_idx = idx + 1

    if start_idx < len(df):
        block = df.iloc[start_idx:]
        if not block.dropna(how="all").empty:
            blocks.append((start_idx, block.reset_index(drop=True)))

    return blocks  # list of (start_row_index, df)


def process_table(df_partial, user_query):
    df_partial = df_partial.dropna(axis=1, how="all")
    raw_headers = df_partial.iloc[0].fillna("Unnamed").astype(str).str.strip()
    df_partial.columns = raw_headers
    df_partial = df_partial[1:].reset_index(drop=True)

    # Fix top-left header if empty
    if df_partial.columns[0].lower() in ["", "unnamed", "nan", "none"]:
        df_partial.columns = ["RowHeader"] + list(df_partial.columns[1:])

    row_header = df_partial.columns[0]
    df_long = df_partial.melt(id_vars=[row_header], var_name="ColumnHeader", value_name="Value")
    df_long.rename(columns={row_header: "RowHeader"}, inplace=True)

    sample = df_long.head(5)
    available_tables = """
    sales_category_gender_region: [Gender Category, Region, Product Category, Sales]
    region_quarter_category_sales: [Region, Quarter, Product Category, Sales]
    """
    prompt = f"""
You are a smart assistant that maps Excel structures to database tables.

User Query:
{user_query}

Excel DataFrame (melted format):
{sample.to_csv(index=False)}

Available tables:
{available_tables}

Return JSON in this format:
{{
  "table": "...",
  "row_header_column": "...",
  "column_header_column": "...",
  "value_column": "...",
  "filters": {{ optional key-value filters like "Product Category": "Eyewear", "Quarter": "Q1" }}
}}
    """

    with st.spinner("Sending prompt to Gemini..."):
        response = model.generate_content(prompt)

    try:
        cleaned_json = re.sub(r"^```json|```$", "", response.text.strip(), flags=re.MULTILINE).strip()
        mapping = json.loads(cleaned_json)
        st.info(f"Using Supabase table: `{mapping['table']}`")
        st.json(mapping)

        # 🧠 Step 4A: Validate column names
        table_meta = supabase.table(mapping["table"]).select("*").limit(1).execute()
        if not table_meta.data:
            st.error("Unable to fetch schema from Supabase.")
            return None
        actual_columns = list(table_meta.data[0].keys())

        for key in ["row_header_column", "column_header_column", "value_column"]:
            current = mapping.get(key, "")
            if current not in actual_columns:
                suggestion = suggest_column_name(current, actual_columns)
                if suggestion:
                    st.warning(f"⚠️ `{current}` not found. Suggesting closest match.")
                    corrected = st.selectbox(
                        f"Replace `{current}` with:",
                        options=[suggestion] + actual_columns,
                        index=0,
                        key=f"fix_{key}"
                    )
                    mapping[key] = corrected
                else:
                    st.error(f"❌ Column `{current}` is invalid and no suggestions found.")
                    return None

    except Exception:
        st.error("Gemini returned invalid JSON. Please check prompt.")
        return None

    # ✅ Step 4B & 4C combined into fetch_value_safe_enhanced
    def fetch_value_safe_enhanced(row_val, col_val, table_name, row_column, col_column, value_column, filters):
        from difflib import get_close_matches

        def get_suggestion(value, options):
            match = get_close_matches(str(value), options, n=1, cutoff=0.6)
            return match[0] if match else None

        try:
            query = (
                supabase.table(table_name)
                .select(value_column)
                .ilike(row_column, str(row_val).strip())
                .ilike(col_column, str(col_val).strip())
            )
            for k, v in filters.items():
                query = query.ilike(k, str(v).strip())
            res = query.execute()

            if not res.data:
                st.warning(f"🔍 No match for `{row_val}, {col_val}`. Trying suggestions...")
                all_row_vals = supabase.table(table_name).select(row_column).execute()
                all_col_vals = supabase.table(table_name).select(col_column).execute()
                row_options = [r[row_column] for r in all_row_vals.data if row_column in r]
                col_options = [r[col_column] for r in all_col_vals.data if col_column in r]

                suggested_row = get_suggestion(row_val, row_options)
                suggested_col = get_suggestion(col_val, col_options)

                if suggested_row:
                    row_val = st.selectbox(f"Did you mean (row)?", [suggested_row] + row_options, key=f"suggest_row_{row_val}")
                if suggested_col:
                    col_val = st.selectbox(f"Did you mean (column)?", [suggested_col] + col_options, key=f"suggest_col_{col_val}")

                query = (
                    supabase.table(table_name)
                    .select(value_column)
                    .ilike(row_column, str(row_val).strip())
                    .ilike(col_column, str(col_val).strip())
                )
                for k, v in filters.items():
                    query = query.ilike(k, str(v).strip())
                res = query.execute()

            # 🔁 Step 4C: Try filter value suggestions
            if not res.data and filters:
                for k, v in filters.items():
                    filter_vals = supabase.table(table_name).select(k).execute()
                    valid_vals = [r[k] for r in filter_vals.data if k in r]
                    suggestion = get_suggestion(v, valid_vals)
                    if suggestion:
                        filters[k] = st.selectbox(f"Did you mean for `{k}`?", [suggestion] + valid_vals, key=f"suggest_filter_{k}_{v}")

                query = (
                    supabase.table(table_name)
                    .select(value_column)
                    .ilike(row_column, str(row_val).strip())
                    .ilike(col_column, str(col_val).strip())
                )
                for k, v in filters.items():
                    query = query.ilike(k, str(v).strip())
                res = query.execute()

            if res.data:
                return sum([r[value_column] for r in res.data])

        except Exception as e:
            st.error(f"❌ Supabase query failed: {e}")
        return None

    # Apply with fallback logic
    df_long[mapping["value_column"]] = df_long.apply(
        lambda row: fetch_value_safe_enhanced(
            row["RowHeader"],
            row["ColumnHeader"],
            mapping["table"],
            mapping["row_header_column"],
            mapping["column_header_column"],
            mapping["value_column"],
            mapping.get("filters", {})
        ), axis=1
    )

    return df_long.pivot(index="RowHeader", columns="ColumnHeader", values=mapping["value_column"]).reset_index()


# --- Excel download ---
def to_excel_download(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()

# --- UI ---
st.title("Enhancement for Waqt")

uploaded_file = st.file_uploader("📤 Upload your Excel file", type=["xlsx"])

if uploaded_file:
    sheets = pd.read_excel(uploaded_file, sheet_name=None, header=None)
    sheet_names = list(sheets.keys())
    selected_sheet = st.selectbox("🧾 Select a sheet to process", sheet_names)
    df = sheets[selected_sheet]

    if df.empty:
        st.warning("❗Selected sheet is empty.")
        st.stop()

    tables = split_dataframe_by_blank_rows(df)

    if not tables:
        st.warning("❗No tables detected in the sheet.")
        st.stop()
    
    # One prompt bar for all tables (split by `;`)
    user_prompt_input = st.text_input(
        "💬 Enter prompts for all tables (separated by `;`)", 
        placeholder="e.g. Sales by Region; Sales by Gender; Sales by Type"
    )
    prompts = [p.strip() for p in user_prompt_input.split(";") if p.strip()]
    positions = [start_row for (start_row, _) in tables]
    
    if len(prompts) != len(tables):
        st.warning(f"⚠️ You entered {len(prompts)} prompts for {len(tables)} tables. Please match the count.")
    else:
        st.markdown("### 🔗 Prompt Mappings")
        for i, prompt in enumerate(prompts):
            st.markdown(f"**Prompt {i+1} → Table {i+1}:** `{prompt}`")

    # 🔁 Process tables on button click
    if len(prompts) == len(tables):
        if st.button("🚀 Start Update"):
            results = []

            for i, ((start_row, table), prompt) in enumerate(zip(tables, prompts)):
                with st.spinner(f"⚙️ Processing Table {i+1}..."):
                    updated = process_table(table, prompt)
                    if updated is not None:
                        results.append((start_row, updated))
                        st.success(f"✅ Table {i+1} updated successfully")
                        st.dataframe(updated)
                    else:
                        st.error(f"❌ Table {i+1} could not be updated.")

            if results:
                wb = load_workbook(uploaded_file)
                ws = wb[selected_sheet]

                for (start_row, updated_df) in results:
                    for r_idx, row in enumerate(dataframe_to_rows(updated_df, index=False, header=True)):
                        for c_idx, value in enumerate(row):
                            ws.cell(row=start_row + r_idx + 1, column=c_idx + 1, value=value)

                output = BytesIO()
                wb.save(output)

                st.download_button(
                    "📥 Download Updated Excel (Original Layout)",
                    data=output.getvalue(),
                    file_name="updated_template_style.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
