import streamlit as st
import pandas as pd
from io import BytesIO
from supabase import create_client
import google.generativeai as genai
import json
import re
from langchain.embeddings import GoogleGenerativeAIEmbeddings
from langchain.vectorstores import FAISS
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.document_loaders import TextLoader
import os


# --- Config ---
st.set_page_config(page_title="Excel Auto-Updater for Waqt", layout="wide")

# --- Environment Secrets ---
SUPABASE_URL = st.secrets.get("SUPABASE_URL")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY")
GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY")

if not SUPABASE_URL or not SUPABASE_KEY or not GEMINI_API_KEY:
    st.error("Missing Supabase or Gemini credentials.")
    st.stop()

# --- Init Clients ---
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel("gemini-2.0-flash")

# --- Title ---
st.title("Excel Auto-Updater for Waqt")


def load_knowledge_base():
    splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
    docs = []

    for filename in os.listdir("knowledge_base"):
        if filename.endswith(".md"):
            loader = TextLoader(os.path.join("knowledge_base", filename))
            docs.extend(splitter.split_documents(loader.load()))

    embedding = GoogleGenerativeAIEmbeddings(model="models/embedding-001")
    vectorstore = FAISS.from_documents(docs, embedding)
    return vectorstore.as_retriever()

def retrieve_rules(prompt_text: str, retriever):
    retrieved_docs = retriever.get_relevant_documents(prompt_text)
    retrieved_knowledge = "\n\n".join([doc.page_content for doc in retrieved_docs])
    
    # Fallback if context is too weak
    if len(retrieved_knowledge.strip()) < 150:
        st.warning("⚠️ Not enough relevant context found. Please rephrase your query.")
        st.stop()

    return retrieved_knowledge
    
# --- Helper Functions ---
def split_dataframe_by_blank_rows(df):
    split_indices = df[df.isnull().all(axis=1)].index.tolist()
    blocks = []
    start_idx = 0
    for idx in split_indices:
        block = df.iloc[start_idx:idx]
        if not block.dropna(how="all").empty:
            blocks.append((start_idx, block.reset_index(drop=True)))
        start_idx = idx + 1
    if start_idx < len(df):
        block = df.iloc[start_idx:]
        if not block.dropna(how="all").empty:
            blocks.append((start_idx, block.reset_index(drop=True)))
    return blocks


def process_table(df_partial_raw):
    df_partial_raw = df_partial_raw.dropna(axis=1, how="all")
    headers = df_partial_raw.iloc[0].fillna("Unnamed").astype(str).str.strip()
    df = df_partial_raw[1:].copy().reset_index(drop=True)
    df.columns = headers
    if df.columns[0].lower() in ["", "unnamed", "nan", "none"]:
        df.columns.values[0] = "RowHeader"
    row_header = df.columns[0]
    df_long = df.melt(id_vars=[row_header], var_name="ColumnHeader", value_name="Value")
    df_long.rename(columns={row_header: "RowHeader"}, inplace=True)
    return df, df_long

# --- Upload File ---
uploaded_file = st.file_uploader("Step 1: Upload your Excel file", type=["xlsx"])

if uploaded_file:
    sheets = pd.read_excel(uploaded_file, sheet_name=None)
    sheet_names = list(sheets.keys())
    selected_sheet = st.selectbox("Select a sheet to process", sheet_names)
    df_raw = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None)

    table_blocks = split_dataframe_by_blank_rows(df_raw)
    table_dfs = []
    orig_headers_list = []
    final_outputs = []
    
    st.subheader(f"📄 Uploaded Template - `{selected_sheet}`")
    for idx, (start_row, block) in enumerate(table_blocks, start=1):
        df_clean, df_long = process_table(block)
        table_dfs.append((df_clean, df_long))
        st.markdown(f"### 🔸 Table {idx} (rows {start_row}–{start_row + len(block) - 1})")
        st.dataframe(df_clean.head(10), use_container_width=True)

    
    # User prompts input
    user_query = st.text_input(
        "Step 2: Enter one prompt per table (separated by `;`)",
        placeholder="e.g. Show average sales by region; Show revenue by gender"
    )

    prompts = [p.strip() for p in user_query.split(";") if p.strip()]

    if user_query and st.button("Start"):
        if len(prompts) != len(table_blocks):
            st.error(f"🚩 You entered {len(prompts)} prompt(s) for {len(table_blocks)} table(s). Please match the count.")
            st.stop()
        retriever = load_knowledge_base()

        for i, ((start_row, raw_block), prompt_text) in enumerate(zip(table_blocks, prompts), start=1):
            df_clean, df_long = process_table(raw_block)
            table_dfs.append((df_clean, df_long))

            orig_headers_list.append(df_clean.columns.tolist())
            
            st.subheader(f"🔹 Preview: Table {i} from {selected_sheet} (rows {start_row}-{start_row + len(raw_block)-1})")
            st.dataframe(df_clean.head(10), use_container_width=True)

            column_info = {
                "productgroup": "Brand of the product (e.g., AI, RG, TF etc.)",
                "product_gender": "Gender the product is designed for (G - Gents, L - Ladies, U - Unisex, P - Pair)",
                "cluster": "Cluster code for internal brand groupings (e.g., LRAGA, LWKWR, GCLSQ etc.)",
                "quantity": "Units sold in the transaction (integer)",
                "billdate": "Date of transaction",
                "channel": "Sales channel (e.g., 1_TW, 2_FASTRACK, 4_MP, 6_HELIOS etc.)",
                "region": "Geographic region (e.g., North, East, South1, West etc.)",
                "raw_region": "Region used specifically for TW, Fastrack, and Helios analysis",
                "tier": "City classification based on business priority (e.g., Metro, Tier 1, Tier 2 etc.)",
                "financial_year": "Financial year of the transaction (e.g., FY23-24)",
                "month_fy": "Month with fiscal year (e.g., Apr FY2425)",
                "value": "Total transaction revenue in INR (numeric)",
                "itemnumber": "Unique SKU or item code",
                "latest_sku": "Parent SKU identifier for grouping variants",
                "ucp_final": "Unit consumer price (selling price per item, numeric)",
                "dealer_type": "Dealer classification for MBR (e.g., EMM, KAM)",
                "platform": "Marketplace platform name (e.g., Amazon, Flipkart)",
                "uid": "customer ID",
                "product_segment": "Product segment (e.g., Smart, Premium, Mainline Analog)",
                "bill_number": "Unique bill or invoice number",
                "store_code": "Internal store identifier",
                "city": "City where the transaction occurred",
                "lfs_chain": "Chain code under LFS channel (e.g., SS, LS etc.)",
                "rs_or_dd": "Dealer model type (RS, DD)",
                "state": "State where the transaction occurred",
                "ytd_tag": "Year-to-date tag for most recent transaction date",
                "dob": "Customer's date of birth",
                "anniversary": "Customer's anniversary date",
                "bday_trans": "Was transaction during customer's birthday window? (Y/N)",
                "anniv_trans": "Was transaction during anniversary window? (Y/N)",
                "customer_gender": "Customer's gender (e.g., Male, Female, Other)"
            }
            
            column_description_text = "\n".join([f"- {k}: {v}" for k, v in column_info.items()])

                     
            # Ensure headers are clean
            headers = df_clean.columns.str.strip().str.replace(" ", "_")
            df_clean.columns = headers
            df_clean = df_clean.fillna("")
            for col in df_clean.select_dtypes(include="object").columns:
                df_clean[col] = df_clean[col].astype(str).str.title()

            row_header = df_clean.columns[0]
            df_long = df_clean.melt(id_vars=[row_header], var_name="ColumnHeader", value_name="Value")
            df_long.rename(columns={row_header: "RowHeader"}, inplace=True)

            sample_rows = []
            for r in df_long["RowHeader"].unique():
                for c in df_long["ColumnHeader"].unique():
                    m = df_long[(df_long["RowHeader"] == r) & (df_long["ColumnHeader"] == c)]
                    if not m.empty:
                        sample_rows.append(m.head(1))
            balanced_sample_df = pd.concat(sample_rows)
            sample_json = json.dumps(balanced_sample_df.to_dict(orient="records"), indent=2)

            retrieved_knowledge = retrieve_rules(prompt_text, retriever)
            
            prompt = f"""
            You are a PostgreSQL expert.
            
            The user has uploaded an Excel sheet that was converted to a long-form JSON structure where:
            - `RowHeader` contains values from one categorical field (e.g., region, gender, productgroup)
            - `ColumnHeader` contains values from another categorical field (e.g., channel, segment, etc.)
            - `Value` is empty and needs to be calculated
            
            ### Context from Knowledge Base:
            {retrieved_knowledge}
            
            ### User Query:
            {prompt_text}
            
            ### Excel JSON Preview:
            {sample_json}
            
            ### Schema:
            {column_description_text}
            
            ### Instructions:
            - Use table: watches_schema
            - Apply filters only using the visible RowHeader and ColumnHeader values
            - Use SUM(), AVG(), COUNT() as needed
            - Return only SQL with three columns: RowHeader, ColumnHeader, AggregatedValue
            - No explanation. No markdown.
            """


            with st.spinner("Sending to Gemini..."):
                response = model.generate_content(prompt)
            sql_query = response.text.strip().strip("`").strip()
            if sql_query.lower().startswith("sql"):
                sql_query = sql_query[3:].strip()
            sql_query = sql_query.rstrip(";")

            with st.expander("Generated SQL Query"):
                st.code(sql_query, language="sql")

            try:
                result = supabase.rpc("run_sql", {"query": sql_query}).execute()
                raw_data = result.data
                if isinstance(raw_data, list) and raw_data and "result" in raw_data[0]:
                    df_result = pd.DataFrame(raw_data[0]["result"])
                else:
                    df_result = pd.DataFrame(raw_data)
            except Exception as e:
                st.error(f"SQL execution failed: {e}")
                continue

            if df_result.empty:
                st.warning("No matching data found.")
                continue

            if df_result.shape[1] == 3:
                # grab the headers the user originally uploaded for this table
                orig = orig_headers_list[i-1]           # e.g. ["RowHeader","Smart","Premium"]
            
                # pivot the returned SQL
                piv = df_result.pivot(
                    index=df_result.columns[0],
                    columns=df_result.columns[1],
                    values=df_result.columns[2]
                )
                # drop any accidental duplicates
                piv = piv.loc[:, ~piv.columns.duplicated()]
            
                # ensure every original column is present
                expected = orig[1:]                     # ["Smart","Premium"]
                for col in expected:
                    if col not in piv.columns:
                        piv[col] = 0                    # or '' if you prefer blanks
            
                # reorder to match the template
                piv = piv[expected]
            
                # reset index and restore original column names
                final_df = piv.reset_index()
                final_df.columns = orig
            else:
                final_df = df_result

            final_outputs.append(final_df)

            st.subheader("Updated Excel Output")
            st.dataframe(final_df, use_container_width=True)

            def to_excel_download(df):
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                return buf.getvalue()

            st.download_button(
                label=f"Download Updated Table {i}",
                data=to_excel_download(final_df),
                file_name=f"updated_table_{i}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Combine all final tables with blank rows in between
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # --- write back into the original template ---
        wb = load_workbook(uploaded_file)
        ws = wb[selected_sheet]
        
        for (start_row, _), final_df in zip(table_blocks, final_outputs):
            # write header + data
            for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True)):
                for c_idx, val in enumerate(row):
                    # +1 because openpyxl is 1-based, and start_row is 0-based
                    ws.cell(row=start_row + r_idx + 1, column=c_idx + 1, value=val)
        
        # dump to bytes and offer download
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        preview_df = pd.read_excel(buf, sheet_name=selected_sheet, header=None)
        st.subheader("🔍 Preview: Updated Template")
        st.dataframe(preview_df, use_container_width=True)
        
        st.download_button(
            "⬇️ Download Updated Excel (Original Layout)",
            data=buf.getvalue(),
            file_name=f"updated_{selected_sheet}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
